
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
import openpyxl
from selenium.common.exceptions import NoSuchElementException

print("Sample test case started")
driver_path = "chromedriver.exe"
path = 'options.xlsx'

# Create a new instance of the Chrome driver
driver = webdriver.Chrome(driver_path)
driver.get("https://www.google.com/")
time.sleep(3)

driver.maximize_window()

# Define the writeData function
def writeData(file_path, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(file_path)

# Load the Excel file
workbook = openpyxl.load_workbook(path)



# Iterate through each day of the week (Saturday to Friday)
for day in workbook.sheetnames:
    print(f"Processing {day}...")

    # Get the corresponding sheet for the current day
    sheet = workbook[day]

    # Get the number of rows in the sheet
    rows = sheet.max_row

    # Iterate over each row in the sheet
    for r in range(3, rows + 1):
        # Read the search query from the Excel file
        search_query = sheet.cell(row=r, column=3).value

        # Find the search box element and enter the search query
        search_box = driver.find_element(By.NAME, "q")

        search_box.clear()
        search_box.send_keys(search_query)

        try:
            # Find the search box element
            search_box = driver.find_element(By.NAME, "q")

            # Enter a partial query to trigger the display of suggested options
            search_box.send_keys("")

            # Wait for the dropdown with suggested options to appear (adjust the wait time as needed)
            time.sleep(3)

            # Find all the options in the dropdown
            options = driver.find_elements(By.XPATH, "//ul[@role='listbox']//li[@class='sbct']")

            # Initialize variables to track the minimum and maximum lengths, along with their corresponding options
            min_length = float('inf')
            min_length_option = None
            max_length = 0
            max_length_option = None

            # Iterate over the options and calculate the lengths
            for option in options:
                option_text = option.text
                option_length = len(option_text)

                # Update the minimum length and the corresponding option if necessary
                if option_length < min_length:
                    min_length = option_length
                    min_length_option = option_text

                # Update the maximum length and the corresponding option if necessary
                if option_length > max_length:
                    max_length = option_length
                    max_length_option = option_text

            # Print the minimum length and the value of the option with the minimum length
            print("Minimum length:", min_length)
            print("Option with minimum length:", min_length_option)

            # Print the maximum length and the value of the option with the maximum length
            print("Maximum length:", max_length)
            print("Option with maximum length:", max_length_option)

            # Write the maximum and minimum values to the Excel file
            writeData(path, day, r, 4, max_length_option)
            writeData(path, day, r, 5, min_length_option)

        except NoSuchElementException:
            print("Element not found")

        search_box.clear()

    print(f"Finished processing {day}\n")

# Close the browser
driver.quit()